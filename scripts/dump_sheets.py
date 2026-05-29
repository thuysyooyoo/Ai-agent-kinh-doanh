# -*- coding: utf-8 -*-
import openpyxl
import sys
import os

wb = openpyxl.load_workbook(r'D:\AI AGENT THUY\AI agent Kinh doanh\templates\Ước tính giá cước ủy thác.xlsx', data_only=False)

output_path = r'D:\AI AGENT THUY\AI agent Kinh doanh\scratch\sheet_structures.txt'
os.makedirs(os.path.dirname(output_path), exist_ok=True)

with open(output_path, 'w', encoding='utf-8') as f:
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        f.write(f"\n================================================================================\n")
        f.write(f"SHEET: {sheet_name} (Max row: {ws.max_row}, Max col: {ws.max_column})\n")
        f.write(f"================================================================================\n")
        
        for r in range(1, ws.max_row + 1):
            row_vals = []
            has_val = False
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                val = cell.value
                if val is not None:
                    has_val = True
                    row_vals.append(f"Col {c}({cell.coordinate}): {repr(val)}")
                else:
                    row_vals.append(f"Col {c}: None")
            if has_val:
                while row_vals and row_vals[-1].endswith(": None"):
                    row_vals.pop()
                non_empty = [v for v in row_vals if not v.endswith(": None")]
                f.write(f"Row {r:03d}: " + " | ".join(non_empty) + "\n")

print("Done! Sheet structures written to scratch/sheet_structures.txt")
