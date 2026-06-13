# -*- coding: utf-8 -*-
import os
import json
import argparse
import sys
import io
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Đảm bảo in ký tự Unicode tiếng Việt không bị lỗi trên Windows console
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

# Đường dẫn mặc định đến file Excel
DEFAULT_EXCEL_PATH = r"d:\AI AGENT THUY\AI agent Kinh doanh\reports\Danh_sach_Van_ban_Phap_luat_XNK.xlsx"

COLUMNS = [
    "STT",
    "Mã Số Văn Bản",
    "Tên Văn Bản",
    "Loại Văn Bản",
    "Cơ Quan Ban Hành",
    "Ngày Ban Hành",
    "Ngày Có Hiệu Lực",
    "Ngày Hết Hiệu Lực",
    "Văn Bản Bị Thay Thế",
    "Phân Tích Sơ Bộ",
    "So Sánh Thay Đổi",
    "Ngày Cập Nhật"
]

def parse_date(date_str):
    """Hỗ trợ phân tích nhiều định dạng ngày phổ biến"""
    if not date_str or pd.isna(date_str) or str(date_str).strip() in ["", "None", "Chưa xác định", "-", "N/A"]:
        return None
    
    date_str = str(date_str).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
    return None

def format_excel_file(file_path):
    """Định dạng Excel đẹp, chuyên nghiệp bằng openpyxl"""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    # Kích hoạt hiển thị đường lưới (grid lines)
    ws.views.sheetView[0].showGridLines = True
    
    # Định nghĩa font chữ và màu sắc
    font_name = "Segoe UI"
    header_font = Font(name=font_name, size=11, bold=True, color="FFFFFF")
    body_font = Font(name=font_name, size=10)
    bold_body_font = Font(name=font_name, size=10, bold=True)
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Xanh navy đậm
    zebra_fill = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")  # Xám xanh nhạt
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    # Định dạng Header
    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    
    ws.row_dimensions[1].height = 28
    
    # Định dạng các dòng dữ liệu
    max_row = ws.max_row
    for row in range(2, max_row + 1):
        # Thiết lập STT tăng dần tự động
        ws.cell(row=row, column=1, value=row - 1)
        
        is_zebra = (row % 2 == 1)
        ws.row_dimensions[row].height = 24
        
        for col in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = body_font
            cell.border = thin_border
            
            if is_zebra:
                cell.fill = zebra_fill
            
            # Căn lề cho các cột
            if col in [1, 2, 4, 6, 7, 8, 9, 12]:  # STT, Mã số, Loại, Các ngày, Văn bản bị thay thế
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:  # Tên văn bản, Cơ quan, Phân tích, So sánh
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                
            # Đắp đậm mã số văn bản
            if col == 2:
                cell.font = bold_body_font
                
    # Thiết lập độ rộng cột tối ưu
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            val_str = str(cell.value or '')
            if cell.row == 1:
                # Tiêu đề cột
                max_len = max(max_len, len(val_str) + 4)
            else:
                max_len = max(max_len, len(val_str))
        
        # Giới hạn độ rộng của các cột chứa nhiều text
        col_idx = col[0].column
        col_name = COLUMNS[col_idx - 1]
        
        if col_name in ["Tên Văn Bản"]:
            ws.column_dimensions[col_letter].width = 35
        elif col_name in ["Phân Tích Sơ Bộ", "So Sánh Thay Đổi"]:
            ws.column_dimensions[col_letter].width = 50
        elif col_name in ["Cơ Cơ Quan Ban Hành"]:
            ws.column_dimensions[col_letter].width = 20
        else:
            # Các cột ngắn tự động căn theo nội dung
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 25)
            
    wb.save(file_path)

def process_tracker(json_data, excel_path):
    print(f"Bắt đầu xử lý tệp Excel tại: {excel_path}")
    
    # 1. Tải file Excel cũ hoặc khởi tạo mới
    if os.path.exists(excel_path):
        try:
            df = pd.read_excel(excel_path)
            # Đồng bộ cột
            for col in COLUMNS:
                if col not in df.columns:
                    df[col] = None
            df = df[COLUMNS]
        except Exception as e:
            print(f"Lỗi khi đọc file Excel hiện tại: {e}. Tiến hành tạo file mới.")
            df = pd.DataFrame(columns=COLUMNS)
    else:
        df = pd.DataFrame(columns=COLUMNS)
    
    # Đảm bảo các kiểu dữ liệu dạng chuỗi để xử lý chính xác
    df["Mã Số Văn Bản"] = df["Mã Số Văn Bản"].astype(str).str.strip()
    
    today = datetime.now().date()
    
    added_count = 0
    skipped_count = 0
    replaced_codes = set()
    
    # 2. Xử lý các văn bản mới
    new_rows = []
    for doc in json_data:
        ma_so = str(doc.get("ma_so", "")).strip()
        if not ma_so:
            continue
            
        # Kiểm tra trùng lặp mã số văn bản
        if ma_so in df["Mã Số Văn Bản"].values:
            print(f"⚠️ Trùng lặp: Văn bản {ma_so} đã tồn tại trong file Excel. Bỏ qua.")
            skipped_count += 1
            continue
            
        ten = doc.get("ten", "")
        loai = doc.get("loai", "")
        co_quan = doc.get("co_quan", "")
        ngay_ban_hanh = doc.get("ngay_ban_hanh", "")
        ngay_hieu_luc = doc.get("ngay_hieu_luc", "")
        ngay_het_hieu_luc = doc.get("ngay_het_hieu_luc", "")
        van_ban_bi_thay_the = str(doc.get("van_ban_bi_thay_the", "")).strip()
        phan_tich = doc.get("phan_tich", "")
        so_sanh = doc.get("so_sanh", "")
        
        # Ghi nhận mã văn bản bị thay thế để xóa sau
        if van_ban_bi_thay_the and van_ban_bi_thay_the not in ["", "None", "N/A", "-"]:
            replaced_codes.add(van_ban_bi_thay_the)
            
        new_row = {
            "STT": None,
            "Mã Số Văn Bản": ma_so,
            "Tên Văn Bản": ten,
            "Loại Văn Bản": loai,
            "Cơ Quan Ban Hành": co_quan,
            "Ngày Ban Hành": ngay_ban_hanh,
            "Ngày Có Hiệu Lực": ngay_hieu_luc,
            "Ngày Hết Hiệu Lực": ngay_het_hieu_luc if ngay_het_hieu_luc else "Chưa xác định",
            "Văn Bản Bị Thay Thế": van_ban_bi_thay_the if van_ban_bi_thay_the else "-",
            "Phân Tích Sơ Bộ": phan_tich,
            "So Sánh Thay Đổi": so_sanh,
            "Ngày Cập Nhật": today.strftime("%Y-%m-%d")
        }
        new_rows.append(new_row)
        added_count += 1
        
    if new_rows:
        df = pd.concat([df, pd.DataFrame(new_rows)], ignore_index=True)
        
    # 3. Thực hiện tự động xóa các văn bản hết hiệu lực
    initial_row_count = len(df)
    rows_to_keep = []
    deleted_expired_count = 0
    deleted_replaced_count = 0
    
    for idx, row in df.iterrows():
        ma_so = str(row["Mã Số Văn Bản"]).strip()
        ngay_het_hieu_luc_str = row["Ngày Hết Hiệu Lực"]
        
        # A. Kiểm tra nếu nằm trong danh sách bị thay thế bởi văn bản mới
        if ma_so in replaced_codes:
            print(f"❌ Xóa văn bản {ma_so} do bị thay thế bởi văn bản mới nhập.")
            deleted_replaced_count += 1
            continue
            
        # B. Kiểm tra nếu ngày hết hiệu lực đã qua
        ngay_het_hieu_luc = parse_date(ngay_het_hieu_luc_str)
        if ngay_het_hieu_luc and ngay_het_hieu_luc <= today:
            print(f"❌ Xóa văn bản {ma_so} do đã hết hiệu lực từ ngày {ngay_het_hieu_luc_str}.")
            deleted_expired_count += 1
            continue
            
        rows_to_keep.append(row)
        
    df = pd.DataFrame(rows_to_keep, columns=COLUMNS)
    
    # Thiết lập lại STT và lưu file
    df["STT"] = range(1, len(df) + 1)
    
    # Đảm bảo thư mục reports tồn tại
    os.makedirs(os.path.dirname(excel_path), exist_ok=True)
    
    # Ghi ra tệp Excel
    df.to_excel(excel_path, index=False)
    
    # Định dạng lại Excel cho chuyên nghiệp
    format_excel_file(excel_path)
    
    print("\n--- BÁO CÁO CẬP NHẬT PHÁP LUẬT ---")
    print(f"📍 Tổng số văn bản mới thêm vào: {added_count}")
    print(f"📍 Số văn bản bị bỏ qua do trùng lặp: {skipped_count}")
    print(f"📍 Số văn bản bị xóa do hết hiệu lực ngày: {deleted_expired_count}")
    print(f"📍 Số văn bản bị xóa do bị thay thế: {deleted_replaced_count}")
    print(f"📍 Tổng số dòng dữ liệu hiện tại trong file: {len(df)}")
    print(f"💾 File Excel đã được lưu tại: {excel_path}")
    print("----------------------------------")

def main():
    parser = argparse.ArgumentParser(description="Quản lý và cập nhật tệp Excel theo dõi văn bản pháp luật XNK.")
    parser.add_argument("-i", "--input", required=True, help="Đường dẫn đến file JSON chứa danh sách văn bản mới.")
    parser.add_argument("-o", "--output", default=DEFAULT_EXCEL_PATH, help="Đường dẫn lưu file Excel kết quả.")
    args = parser.parse_args()
    
    if not os.path.exists(args.input):
        print(f"Lỗi: Không tìm thấy file JSON dữ liệu đầu vào tại: {args.input}")
        return
        
    try:
        with open(args.input, "r", encoding="utf-8") as f:
            json_data = json.load(f)
    except Exception as e:
        print(f"Lỗi khi đọc file JSON đầu vào: {e}")
        return
        
    process_tracker(json_data, args.output)

if __name__ == "__main__":
    main()
