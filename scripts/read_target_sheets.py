# -*- coding: utf-8 -*-
import openpyxl
import sys
import os

wb_formula = openpyxl.load_workbook(r'D:\AI AGENT THUY\AI agent Kinh doanh\templates\Ước tính giá cước ủy thác.xlsx', data_only=False)
wb_val = openpyxl.load_workbook(r'D:\AI AGENT THUY\AI agent Kinh doanh\templates\Ước tính giá cước ủy thác.xlsx', data_only=True)

output_path = r'D:\AI AGENT THUY\AI agent Kinh doanh\scratch\target_sheets_analysis.txt'

with open(output_path, 'w', encoding='utf-8') as f:
    def write_sheet_comparison(sheet_name):
        ws_f = wb_formula[sheet_name]
        ws_v = wb_val[sheet_name]
        
        f.write(f"\n==================================================================================\n")
        f.write(f"SHEET DETAIL: {sheet_name}\n")
        f.write(f"==================================================================================\n")
        
        for r in range(1, ws_f.max_row + 1):
            row_cells = []
            has_data = False
            for c in range(1, ws_f.max_column + 1):
                cell_f = ws_f.cell(row=r, column=c)
                cell_v = ws_v.cell(row=r, column=c)
                
                val_f = cell_f.value
                val_v = cell_v.value
                
                if val_f is not None:
                    has_data = True
                    coord = cell_f.coordinate
                    if str(val_f).startswith("="):
                        row_cells.append(f"{coord}: [Formula: {val_f} | Value: {val_v}]")
                    else:
                        row_cells.append(f"{coord}: {repr(val_f)}")
            if has_data:
                f.write(f"Row {r:02d}: " + " | ".join(row_cells) + "\n")

    write_sheet_comparison("1 sổ gom cont")
    write_sheet_comparison("1 sổ nguyên lô ")

print("Done writing to UTF-8 file!")
