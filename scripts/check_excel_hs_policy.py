# -*- coding: utf-8 -*-
import os
import sys
import re
import argparse
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

def clean_hs_code(hs):
    if not hs:
        return []
    # Replace common separators with semicolons
    s = str(hs).replace(',', ';').replace('\n', ';').replace('/', ';')
    parts = s.split(';')
    codes = []
    for p in parts:
        cleaned = re.sub(r'[^0-9]', '', p)
        if cleaned:
            codes.append(cleaned)
    return codes

def search_excel_hs(query_hs, excel_path):
    cleaned_query = re.sub(r'[^0-9]', '', query_hs)
    if not cleaned_query:
        return []
        
    if not os.path.exists(excel_path):
        print(f"Error: Excel file not found at {excel_path}")
        sys.exit(1)
        
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    results = []
    
    for sheetname in ["Rủi ro cao", "Rủi ro trung bình"]:
        if sheetname not in wb.sheetnames:
            continue
        ws = wb[sheetname]
        risk_level = "Cao" if "cao" in sheetname.lower() else "Trung bình"
        
        # Read only up to max_row (which openpyxl read_only might not calculate exactly, so we loop safely)
        for r, row in enumerate(ws.iter_rows(min_row=2), start=2):
            # Check row length and unpack
            if len(row) < 7:
                continue
                
            stt = row[0].value
            name = row[1].value
            qcvn = row[2].value
            hs = row[3].value
            desc = row[4].value
            req = row[5].value
            dept = row[6].value
            
            if stt is None and name is None and hs is None and dept is None:
                continue
                
            cleaned_codes = clean_hs_code(hs)
            if not cleaned_codes:
                continue
                
            for code in cleaned_codes:
                # 1. Exact Match
                if cleaned_query == code:
                    results.append({
                        "match_type": "Khớp chính xác",
                        "matched_key": code,
                        "data": {
                            "STT": str(stt) if stt is not None else "",
                            "Tên sản phẩm, hàng hóa": str(name) if name is not None else "",
                            "Quy chuẩn kỹ thuật/Tiêu chuẩn": str(qcvn) if qcvn is not None else "",
                            "Mã số HS gốc": str(hs) if hs is not None else "",
                            "Mô tả sản phẩm, hàng hóa": str(desc) if desc is not None else "",
                            "Yêu cầu quản lý chất lượng tương ứng": str(req) if req is not None else "",
                            "Bộ ban ngành quản lý": str(dept) if dept is not None else "",
                            "Mức độ rủi ro": risk_level
                        }
                    })
                    break
                # 2. Prefix Match: Excel key is a prefix of the query (e.g. key is '8517' and query is '85171100')
                elif cleaned_query.startswith(code):
                    results.append({
                        "match_type": f"Khớp tiền tố (Nhóm {code})",
                        "matched_key": code,
                        "data": {
                            "STT": str(stt) if stt is not None else "",
                            "Tên sản phẩm, hàng hóa": str(name) if name is not None else "",
                            "Quy chuẩn kỹ thuật/Tiêu chuẩn": str(qcvn) if qcvn is not None else "",
                            "Mã số HS gốc": str(hs) if hs is not None else "",
                            "Mô tả sản phẩm, hàng hóa": str(desc) if desc is not None else "",
                            "Yêu cầu quản lý chất lượng tương ứng": str(req) if req is not None else "",
                            "Bộ ban ngành quản lý": str(dept) if dept is not None else "",
                            "Mức độ rủi ro": risk_level
                        }
                    })
                    break
                # 3. Sub-code Match: Query is a prefix of the Excel key (e.g. query is '8517' and key is '85171100')
                elif code.startswith(cleaned_query):
                    results.append({
                        "match_type": f"Khớp mã con (Mã HS {code})",
                        "matched_key": code,
                        "data": {
                            "STT": str(stt) if stt is not None else "",
                            "Tên sản phẩm, hàng hóa": str(name) if name is not None else "",
                            "Quy chuẩn kỹ thuật/Tiêu chuẩn": str(qcvn) if qcvn is not None else "",
                            "Mã số HS gốc": str(hs) if hs is not None else "",
                            "Mô tả sản phẩm, hàng hóa": str(desc) if desc is not None else "",
                            "Yêu cầu quản lý chất lượng tương ứng": str(req) if req is not None else "",
                            "Bộ ban ngành quản lý": str(dept) if dept is not None else "",
                            "Mức độ rủi ro": risk_level
                        }
                    })
                    break
    return results

def main():
    parser = argparse.ArgumentParser(description="Tra cứu chính sách hàng hóa nhập khẩu rủi ro trực tiếp từ file Excel.")
    parser.add_argument("--hs", type=str, required=True, help="Mã HS cần tra cứu (ví dụ: 8507 hoặc 8603)")
    args = parser.parse_args()
    
    excel_path = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        'knowledge', 'customs_rules', 'DANH_MUC_HANG_HOA_RUI_RO_TONG_HOP.xlsx'
    )
    query = args.hs.strip()
    
    results = search_excel_hs(query, excel_path)
    
    if not results:
        print(f"### Kết quả tra cứu HS Code: {query}")
        print("\n❌ Không tìm thấy thông tin rủi ro/tiền kiểm/hậu kiểm cho mã HS này từ file Excel tổng hợp. Hàng hóa có thể thuộc diện hàng thông thường hoặc cần tra cứu thủ công.")
        sys.exit(0)
        
    print(f"### Kết quả tra cứu HS Code (Trực tiếp từ Excel): {query}")
    print(f"Tìm thấy **{len(results)}** kết quả phù hợp:\n")
    
    limit = 15
    for idx, r in enumerate(results[:limit]):
        data = r["data"]
        print(f"#### {idx+1}. {r['match_type']} - Mã gốc trong danh mục: `{data['Mã số HS gốc']}`")
        print(f"- **Mức độ rủi ro:** **{data['Mức độ rủi ro']}**")
        print(f"- **Bộ ban ngành quản lý:** {data['Bộ ban ngành quản lý']}")
        print(f"- **Tên sản phẩm:** {data['Tên sản phẩm, hàng hóa']}")
        print(f"- **Quy chuẩn kỹ thuật/Tiêu chuẩn:** `{data['Quy chuẩn kỹ thuật/Tiêu chuẩn']}`")
        print(f"- **Mô tả:** {data['Mô tả sản phẩm, hàng hóa']}")
        print(f"- **Yêu cầu quản lý chất lượng:** {data['Yêu cầu quản lý chất lượng tương ứng']}")
        print("-" * 40)
        
    if len(results) > limit:
        print(f"\n*(Đã ẩn {len(results) - limit} kết quả do danh sách quá dài. Vui lòng tra cứu cụ thể hơn)*")

if __name__ == "__main__":
    main()
