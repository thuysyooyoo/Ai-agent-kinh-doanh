# -*- coding: utf-8 -*-
import openpyxl
import os
import shutil
import re
import json
import sys

sys.stdout.reconfigure(encoding='utf-8')

def clean_hs_to_digits(hs):
    if not hs:
        return []
    # Replace common delimiters with a semicolon
    s = str(hs).replace(',', ';').replace('\n', ';').replace('/', ';')
    parts = s.split(';')
    codes = []
    for p in parts:
        cleaned = re.sub(r'[^0-9]', '', p)
        if cleaned:
            codes.append(cleaned)
    return codes

def main():
    excel_path = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        'knowledge', 'customs_rules', 'DANH_MUC_HANG_HOA_RUI_RO_TONG_HOP.xlsx'
    )
    backup_path = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        'knowledge', 'customs_rules', 'DANH_MUC_HANG_HOA_RUI_RO_TONG_HOP_BACKUP.xlsx'
    )
    
    script_dir = os.path.dirname(os.path.abspath(__file__))
    json_path = os.path.join(script_dir, "..", "knowledge", "customs_rules", "danh_muc_tra_cuu_hs_code_rui_ro.json")
    json_path = os.path.normpath(json_path)

    # 1. Create backup of Excel file
    if not os.path.exists(backup_path):
        shutil.copy(excel_path, backup_path)
        print(f"[Backup] Created Excel backup at: {backup_path}")
    else:
        print(f"[Backup] Excel backup already exists at: {backup_path}")

    # 2. Load Excel Workbook
    wb = openpyxl.load_workbook(excel_path)
    
    # Track statistics
    stats = {
        "Rủi ro cao": {"BXD_updates": 0, "BCT_updates": 0},
        "Rủi ro trung bình": {"BXD_updates": 0, "BCT_updates": 0}
    }
    
    for sheetname in ["Rủi ro cao", "Rủi ro trung bình"]:
        if sheetname not in wb.sheetnames:
            print(f"[Error] Sheet '{sheetname}' not found in workbook.")
            continue
            
        ws = wb[sheetname]
        print(f"\n[Excel] Processing sheet: {sheetname}")
        
        for r in range(2, ws.max_row + 1):
            stt = ws.cell(row=r, column=1).value
            name = ws.cell(row=r, column=2).value
            qcvn = ws.cell(row=r, column=3).value
            hs = ws.cell(row=r, column=4).value
            desc = ws.cell(row=r, column=5).value
            req_cell = ws.cell(row=r, column=6)
            dept_cell = ws.cell(row=r, column=7)
            
            dept = dept_cell.value
            req = req_cell.value
            hs_str = str(hs) if hs else ""
            
            # GTVT -> BXD update
            if dept == "Bộ Giao thông Vận tải":
                is_railway_or_electric = False
                # Check QCVN for BXD
                if qcvn and "/BXD" in str(qcvn):
                    is_railway_or_electric = True
                # Check railway code (starting with 86)
                elif hs_str.startswith("86") or hs_str.startswith("86."):
                    is_railway_or_electric = True
                    
                if is_railway_or_electric:
                    dept_cell.value = "Bộ Xây dựng"
                    stats[sheetname]["BXD_updates"] += 1
                    
            # BCT placeholder update
            if dept == "Bộ Công Thương":
                if req and "/2026/TT-BCT" in str(req):
                    req_cell.value = str(req).replace("/2026/TT-BCT", "33/2026/TT-BCT")
                    stats[sheetname]["BCT_updates"] += 1
                    
    # Save the updated Excel
    wb.save(excel_path)
    print("\n[Excel] Updates completed and saved successfully:")
    for sheetname, s in stats.items():
        print(f" - {sheetname}: Moved {s['BXD_updates']} items to 'Bộ Xây dựng', Updated {s['BCT_updates']} BCT placeholders.")

    # 3. Reload Excel data-only to export to JSON
    wb_data = openpyxl.load_workbook(excel_path, data_only=True)
    json_db = {}
    
    for sheetname in ["Rủi ro cao", "Rủi ro trung bình"]:
        ws = wb_data[sheetname]
        risk_level = "Cao" if "cao" in sheetname.lower() else "Trung bình"
        
        for r in range(2, ws.max_row + 1):
            stt = ws.cell(row=r, column=1).value
            name = ws.cell(row=r, column=2).value
            qcvn = ws.cell(row=r, column=3).value
            hs = ws.cell(row=r, column=4).value
            desc = ws.cell(row=r, column=5).value
            req = ws.cell(row=r, column=6).value
            dept = ws.cell(row=r, column=7).value
            
            # If all are None, skip the empty row
            if stt is None and name is None and hs is None and dept is None:
                continue
                
            # Clean and map to dictionary
            entry = {
                "STT": str(stt) if stt is not None else "",
                "Tên sản phẩm, hàng hóa": str(name) if name is not None else "",
                "Quy chuẩn kỹ thuật/Tiêu chuẩn": str(qcvn) if qcvn is not None else "",
                "Mã số HS gốc": str(hs) if hs is not None else "",
                "Mô tả sản phẩm, hàng hóa": str(desc) if desc is not None else "",
                "Yêu cầu quản lý chất lượng tương ứng": str(req) if req is not None else "",
                "Bộ ban ngành quản lý": str(dept) if dept is not None else "",
                "Mức độ rủi ro": risk_level
            }
            
            # Clean HS code to use as dictionary keys
            cleaned_codes = clean_hs_to_digits(hs)
            if not cleaned_codes:
                # If there's no HS code listed in the row, we don't index it in JSON (similar to old database)
                continue
                
            for code in cleaned_codes:
                if code not in json_db:
                    json_db[code] = []
                json_db[code].append(entry)

    # 4. Write the JSON database
    os.makedirs(os.path.dirname(json_path), exist_ok=True)
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(json_db, f, ensure_ascii=False, indent=2)
        
    print(f"\n[JSON] Exported {len(json_db)} unique HS code keys to search database: {json_path}")
    print("[JSON] Synchronization complete.")

if __name__ == "__main__":
    main()
