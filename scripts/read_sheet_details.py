# -*- coding: utf-8 -*-
import openpyxl
import sys
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(r'D:\AI AGENT THUY\AI agent Kinh doanh\templates\Ước tính giá cước ủy thác.xlsx', data_only=False)

def inspect_sheet(sheet_name):
    ws = wb[sheet_name]
    print(f"\n=======================================================")
    print(f"SHEET: {sheet_name} (Max row: {ws.max_row}, Max col: {ws.max_column})")
    print(f"=======================================================")
    
    # Read cells
    for r in range(1, min(ws.max_row + 1, 100)):
        row_vals = []
        has_val = False
        for c in range(1, min(ws.max_column + 1, 30)):
            cell = ws.cell(row=r, column=c)
            val = cell.value
            if val is not None:
                has_val = True
                row_vals.append(f"C{c}({cell.coordinate}): {repr(val)}")
            else:
                row_vals.append(f"C{c}: None")
        if has_val:
            # Filter out trailing Nones for cleaner display
            while row_vals and row_vals[-1].endswith(": None"):
                row_vals.pop()
            non_empty_vals = [v for v in row_vals if not v.endswith(": None")]
            print(f"Row {r:02d}: " + " | ".join(non_empty_vals))

inspect_sheet("1 sổ gom cont")
inspect_sheet("1 sổ nguyên lô ")
